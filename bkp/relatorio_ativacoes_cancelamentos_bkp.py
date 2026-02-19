import pandas as pd
import awswrangler as awr
import openpyxl
import shutil
import datetime as dt
import pyautogui
import time
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

class RelatorioAtivacoesCancelamentos:
    def __init__(self):
        print("[1/9] Iniciando RelatorioAtivacoesCancelamentos...")
        self.paths = {
            "ativos_sql": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_ATIVOS.sql",
            "cancelados_integrais_sql": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_INTEGRAIS.sql",
            "cancelados_parciais_sql": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_PARCIAIS.sql",
            "template_xlsx": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\template\TEMPLATE_BASE_ATIVACOES_CANCELAMENTOS.xlsx",
            "img_output": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\img",
            "report_output": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\reports",
            "sharepoint_path": r"C:\Users\raphael.almeida\OneDrive - Grupo Unus\analise de dados - Arquivos em excel",
        }

        print("[2/9] Lendo arquivos .sql...")
        self.ativos_sql = self._read_sql_file(self.paths["ativos_sql"])
        self.cancelados_integrais_sql = self._read_sql_file(self.paths["cancelados_integrais_sql"])
        self.cancelados_parciais_sql = self._read_sql_file(self.paths["cancelados_parciais_sql"])
        self.yesterday, self.yesterday_ts, self.yesterday_date = self._get_yesterday_date()
        self.load_data()
        self.format_all_types()
        self.process_data()
        self.load_workbook_and_sheets()
        print("[3/9] Inicialização completa.\n")

    def _read_sql_file(self, path):
        with open(path, 'r', encoding='utf-8') as file:
            return file.read()
    
    def _get_yesterday_date(self):
        today_ts = pd.Timestamp.today().normalize()
        if today_ts.weekday() == 0:
            yesterday_ts = today_ts - dt.timedelta(days=3)
        else:
            yesterday_ts = today_ts - dt.timedelta(days=1)
        yesterday = yesterday_ts.strftime('%d-%m-%Y')
        return yesterday, yesterday_ts, yesterday_ts.date()

    def load_data(self):
        print("[4/9] Carregando dados do Athena...")
        self.df_ativos = awr.athena.read_sql_query(sql=self.ativos_sql, database='silver').drop(columns=['rn'])
        self.df_ativos = (self.df_ativos
                          .sort_values(by=['chassi', 'inicio_vig', 'data_ativacao'], ascending=[True, False, False])
                          .drop_duplicates(subset=['chassi'], keep='first'))
        self.df_cancelamentos_integrais = awr.athena.read_sql_query(sql=self.cancelados_integrais_sql, database='silver')
        self.df_cancelamentos_parciais = awr.athena.read_sql_query(sql=self.cancelados_parciais_sql, database='silver')
        print("... Dados carregados.")

    def format_type(self, df):
        for col in df.select_dtypes(include=['string']).columns:
            df[col] = df[col].str.upper()
        return df

    def format_all_types(self):
        # Pequeno log sobre normalização
        print("[5/9] Formatando as colunas dos DataFrames para garantir consistência (strings em maiúsculo, NaNs)...")
        self.format_type(self.df_ativos)
        self.format_type(self.df_cancelamentos_integrais)
        self.format_type(self.df_cancelamentos_parciais)

    def process_data(self):
        print("[6/9] Processando e agregando dados de ativações e cancelamentos...")
        # CANCELAMENTOS
        self.df_cancelamentos_integrais['data_cancelamento'] = pd.to_datetime(
            self.df_cancelamentos_integrais['data_cancelamento'], errors='coerce'
        ).dt.date
        self.placas_canceladas_dia_anterior = self.df_cancelamentos_integrais.loc[
            self.df_cancelamentos_integrais['data_cancelamento'] == self.yesterday_date, 'placa'
        ].unique().tolist()
        self.df_cancelamentos_integrais['identificador'] = 'INTEGRAL'
        self.df_cancelamentos_parciais['identificador'] = 'PARCIAL'
        self.df_cancelamentos = pd.concat([
            self.df_cancelamentos_integrais,
            self.df_cancelamentos_parciais
        ], ignore_index=True)
        self.df_cancelamentos = self.df_cancelamentos.sort_values(
            by='data_cancelamento', ascending=False
        ).reset_index(drop=True)
        self.df_cancelamentos = self.df_cancelamentos.drop_duplicates(subset=['chassi', 'empresa', 'coverage_id'], keep='first')
        self.df_ativos = self.df_ativos.sort_values(
            by=['inicio_vig', 'data_ativacao'], ascending=False
        ).reset_index(drop=True)
        # YESTERDAY VARIABLE
        today = dt.date.today()
        if today.weekday() == 0:
            self.yesterday_yyyymmdd = today - dt.timedelta(days=3)
        else:
            self.yesterday_yyyymmdd = today - dt.timedelta(days=1)
        self.yesterday = self.yesterday_yyyymmdd
        # ATIVADOS E CANCELADOS DASH NUMBERS
        self.ativados = {
            'SEGTRUCK': len(self.df_ativos[(self.df_ativos['empresa'] == 'SEGTRUCK') & (self.df_ativos['inicio_vig'] == self.yesterday)]),
            'STCOOP': len(self.df_ativos[(self.df_ativos['empresa'] == 'STCOOP') & (self.df_ativos['inicio_vig'] == self.yesterday)]),
            'VIAVANTE': len(self.df_ativos[(self.df_ativos['empresa'] == 'VIAVANTE') & (self.df_ativos['inicio_vig'] == self.yesterday)]),
            'TAG': len(self.df_ativos[(self.df_ativos['empresa'] == 'TAG') & (self.df_ativos['inicio_vig'] == self.yesterday)]),
        }
        self.cancelados = {
            'SEGTRUCK': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'SEGTRUCK') & (self.df_cancelamentos['data_cancelamento'] == self.yesterday)]),
            'STCOOP': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'STCOOP') & (self.df_cancelamentos['data_cancelamento'] == self.yesterday)]),
            'VIAVANTE': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'VIAVANTE') & (self.df_cancelamentos['data_cancelamento'] == self.yesterday)]),
            'TAG': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'TAG') & (self.df_cancelamentos['data_cancelamento'] == self.yesterday)]),
        }
        print("... Dados processados.")

    def load_workbook_and_sheets(self):
        print("[7/9] Carregando template Excel e preenchendo abas de ATIVOS e CANCELAMENTOS...")
        self.wb = openpyxl.load_workbook(self.paths["template_xlsx"])
        self.ws_ativos = self.wb['ATIVOS']
        self.ws_cancelamentos = self.wb['CANCELAMENTOS']
        self.ws_base = self.wb['BASE']
        self.ws_relatorio = self.wb['RELATORIO']
        self.clear_sheet(self.ws_ativos)
        self.clear_sheet(self.ws_cancelamentos)
        self.fill_nan(self.df_ativos)
        self.fill_nan(self.df_cancelamentos)
        self.write_to_sheet(self.df_ativos, self.ws_ativos)
        self.write_to_sheet(self.df_cancelamentos, self.ws_cancelamentos)
        print("... Abas ATIVOS e CANCELAMENTOS preenchidas.")

    def clear_sheet(self, sheet):
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(1, max_row)

    def fill_nan(self, df):
        numeric_cols = df.select_dtypes(include=['number']).columns
        string_cols = df.select_dtypes(include=['object', 'string']).columns
        df[numeric_cols] = df[numeric_cols].fillna(0)
        df[string_cols] = df[string_cols].fillna('N/A')

    def write_to_sheet(self, df, ws):
        # Cabeçalhos
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        # Dados
        if not df.empty:
            for r_idx, row in enumerate(df.values, start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    def update_base_sheet(self):
        ws = self.ws_base
        df_ativos = self.df_ativos
        yesterday = self.yesterday
        # Encontrar primeira linha vazia na coluna B
        first_empty_row = ws.max_row + 1
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=2).value is None:
                first_empty_row = row
                break
        # Pega data atual
        data_atual_planilha = ws['A' + str(first_empty_row)].value
        # Parse dates
        data_atual_planilha = self._to_date_obj(data_atual_planilha)
        yesterday_date = self._to_date_obj(yesterday)
        # Coloca o registro de ativos se a data coincidir
        if data_atual_planilha == yesterday_date:
            qtd_ativos = 0
            if "data_ativacao" in df_ativos.columns:
                data_ativacao_col = pd.to_datetime(df_ativos["data_ativacao"], errors='coerce')
                df_ativos_filtrados = df_ativos[data_ativacao_col <= pd.to_datetime(yesterday_date)]
                qtd_ativos = df_ativos_filtrados['chassi'].nunique()
            else:
                qtd_ativos = df_ativos['chassi'].nunique()
            dia_semana = yesterday_date.weekday()
            if dia_semana == 5:
                ws['B' + str(first_empty_row)] = 'SÁBADO'
            elif dia_semana == 6:
                ws['B' + str(first_empty_row)] = 'DOMINGO'
            else:
                ws['B' + str(first_empty_row)] = qtd_ativos

    def _to_date_obj(self, dt_like):
        if isinstance(dt_like, pd.Timestamp):
            return dt_like.date()
        elif isinstance(dt_like, dt.datetime):
            return dt_like.date()
        elif hasattr(dt_like, 'date'):
            return dt_like.date()
        elif isinstance(dt_like, str):
            return pd.to_datetime(dt_like).date()
        return dt_like

    def preencher_relatorio(self):
        # Preenche a aba RELATORIO com valores de ativos/cancelados (pequeno log)
        print("[8/9] Preenchendo aba RELATORIO com os valores principais...")
        ws = self.ws_relatorio
        a = self.ativados
        c = self.cancelados
        ws['C6'] = a['SEGTRUCK']
        ws['C7'] = a['STCOOP']
        ws['C8'] = a['VIAVANTE']
        ws['C9'] = a['TAG']
        ws['D6'] = c['SEGTRUCK']
        ws['D7'] = c['STCOOP']
        ws['D8'] = c['VIAVANTE']
        ws['D9'] = c['TAG']

    def criar_imagem_resumo(self):
        print("[9/9] Criando imagem resumo de ativos/cancelamentos...")
        # Totais
        a, c = self.ativados, self.cancelados
        total_ativados = sum(a.values())
        total_cancelados = sum(c.values())
        # Data format
        yesterday = self.yesterday
        if hasattr(yesterday, 'strftime'):
            data_atual = yesterday.strftime('%d/%m/%Y')
            data_antiga = (yesterday - dt.timedelta(days=1)).strftime('%d/%m/%Y')
        else:
            y = str(yesterday)
            parts = y.split('-')
            if len(parts) == 3:
                data_atual = f'{parts[2]}/{parts[1]}/{parts[0]}'
            else:
                data_atual = y
        # Total ativos geral
        if "data_ativacao" in self.df_ativos.columns:
            data_ativacao_col = pd.to_datetime(self.df_ativos["data_ativacao"], errors='coerce')
            df_ativos_filtrados = self.df_ativos[data_ativacao_col <= pd.to_datetime(self._to_date_obj(yesterday))]
            total_ativos_geral = f"{df_ativos_filtrados['chassi'].nunique():,}".replace(',', '.')
        else:
            total_ativos_geral = f"{self.df_ativos['chassi'].nunique():,}".replace(',', '.')
        # Desenhar gráfico
        colunas_main = ['EMPRESA', 'ATIVADOS', 'CANCELADOS']
        dados_main = [
            ['Segtruck', a['SEGTRUCK'], c['SEGTRUCK']],
            ['Stcoop', a['STCOOP'], c['STCOOP']],
            ['Viavante', a['VIAVANTE'], c['VIAVANTE']],
            ['Tag', a['TAG'], c['TAG']],
            ['Total', total_ativados, total_cancelados]
        ]
        dados_topo = [ ['DATA', data_atual], ['TOTAL ATIVOS', total_ativos_geral] ]
        cor_fundo_cinza = '#E6E6E6'
        cor_fundo_branco = '#FFFFFF'

        fig, ax = plt.subplots(figsize=(6, 5))
        ax.axis('off')
        table_top = ax.table(cellText=dados_topo, cellLoc='center', loc='center', bbox=[0.0, 0.75, 1.0, 0.2])
        for (row, col), cell in table_top.get_celld().items():
            cell.set_height(0.1)
            cell.set_text_props(fontweight='bold', color='#000000')
            cell.set_facecolor(cor_fundo_cinza if col == 0 else cor_fundo_branco)
        table_main = ax.table(
            cellText=dados_main, colLabels=colunas_main,
            cellLoc='center', loc='center', bbox=[0.0, 0.0, 1.0, 0.7]
        )
        for (row, col), cell in table_main.get_celld().items():
            cell.set_height(0.12)
            if row == 0: cell.set_facecolor(cor_fundo_cinza); cell.set_text_props(fontweight='bold', color='#000000')
            elif col == 0: cell.set_facecolor(cor_fundo_cinza); cell.set_text_props(fontweight='bold', color='#000000')
            elif row == 3: cell.set_facecolor(cor_fundo_branco); cell.set_text_props(fontweight='normal', color='#000000')
            elif row == 5: cell.set_facecolor(cor_fundo_cinza); cell.set_text_props(fontweight='bold', color='#000000')
            else: cell.set_facecolor(cor_fundo_branco); cell.set_text_props(fontweight='normal', color='#000000')
        plt.tight_layout()
        os.makedirs(self.paths["img_output"], exist_ok=True)
        # Exclui o arquivo correspondente à data_antiga, se existir
        if 'data_antiga' in locals():
            image_path_antiga = os.path.join(
                self.paths["img_output"], f'tabela_ativacoes_cancelamentos_{data_antiga.replace("/", "-")}.png'
            )
            if os.path.exists(image_path_antiga):
                try:
                    os.remove(image_path_antiga)
                except Exception as e:
                    print(f"Erro ao remover arquivo antigo: {image_path_antiga}. Erro: {e}")
        image_path = os.path.join(self.paths["img_output"], f'tabela_ativacoes_cancelamentos_{data_atual.replace("/", "-")}.png')
        plt.savefig(image_path, dpi=300, bbox_inches='tight')
        plt.close()
        print("[OK] Imagem resumo criada.")

    def criar_graficos_unidades(self):
        print("[INFO] Criando gráficos de unidades de ativos/cancelamentos...")
        yesterday = self.yesterday
        if hasattr(yesterday, 'strftime'):
            yesterday_str = yesterday.strftime('%d/%m/%Y')
            dbf_date = yesterday - dt.timedelta(days=1)
            dbf_yesterday_str = dbf_date.strftime('%d/%m/%Y')
        else:
            y = str(yesterday)
            parts = y.split('-')
            if len(parts) == 3:
                yesterday_str = f'{parts[2]}/{parts[1]}/{parts[0]}'
                # dbf = day before
                dbf_parts = [int(p) for p in parts]
                try:
                    dbf_date = dt.date(dbf_parts[0], dbf_parts[1], dbf_parts[2]) - dt.timedelta(days=1)
                    dbf_yesterday_str = dbf_date.strftime('%d/%m/%Y')
                except Exception:
                    dbf_yesterday_str = f'{parts[2]}/{parts[1]}/{parts[0]}'
            else:
                yesterday_str = y
                dbf_yesterday_str = y

        # Ativos por unidade
        df_ativos_ontem = self.df_ativos[self.df_ativos['data_ativacao'] == yesterday]
        df_ativos_ontem_unidades = df_ativos_ontem.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)
        # Cancelamentos por unidade
        df_cancelamentos_ontem = self.df_cancelamentos[self.df_cancelamentos['data_cancelamento'] == yesterday]
        df_cancelamentos_ontem_unidades = df_cancelamentos_ontem.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)
        # Plot
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 6))
        if not df_ativos_ontem_unidades.empty:
            df_ativos_ontem_unidades.plot(kind='barh', ax=ax1, color='#2ecc71', edgecolor='black', linewidth=0.5)
            ax1.set_title(f'Ativos por Unidade - {yesterday_str}', fontsize=14, fontweight='bold', pad=15)
            ax1.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax1.set_ylabel('', fontsize=11, fontweight='bold')
            ax1.set_xticks([])
            ax1.grid(axis='x', alpha=0.3, linestyle='--', visible=False)
            ax1.invert_yaxis()
            xmax = df_ativos_ontem_unidades.values.max() if len(df_ativos_ontem_unidades) > 0 else 1
            ax1.set_xlim(0, xmax + max(3, int(0.15 * xmax)))
            for i, v in enumerate(df_ativos_ontem_unidades.values):
                ax1.text(v + 0.7, i, str(int(v)), va='center', fontweight='bold')
        else:
            ax1.text(0.5, 0.5, 'Sem dados disponíveis', ha='center', va='center', transform=ax1.transAxes, fontsize=12)
            ax1.set_title(f'Ativos por Unidade - {yesterday_str}', fontsize=14, fontweight='bold')
            ax1.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax1.set_ylabel('', fontsize=11, fontweight='bold')
            ax1.set_xticks([])

        if not df_cancelamentos_ontem_unidades.empty:
            df_cancelamentos_ontem_unidades.plot(kind='barh', ax=ax2, color='#e74c3c', edgecolor='black', linewidth=0.5)
            ax2.set_title(f'Cancelados por Unidade - {yesterday_str}', fontsize=14, fontweight='bold', pad=15)
            ax2.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax2.set_ylabel('', fontsize=11, fontweight='bold')
            ax2.set_xticks([])
            ax2.grid(axis='x', alpha=0.3, linestyle='--', visible=False)
            ax2.invert_yaxis()
            xmax2 = df_cancelamentos_ontem_unidades.values.max() if len(df_cancelamentos_ontem_unidades) > 0 else 1
            ax2.set_xlim(0, xmax2 + max(3, int(0.15 * xmax2)))
            for i, v in enumerate(df_cancelamentos_ontem_unidades.values):
                ax2.text(v + 0.7, i, str(int(v)), va='center', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'Sem dados disponíveis', ha='center', va='center', transform=ax2.transAxes, fontsize=12)
            ax2.set_title(f'Cancelados por Unidade - {yesterday_str}', fontsize=14, fontweight='bold')
            ax2.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax2.set_ylabel('', fontsize=11, fontweight='bold')
            ax2.set_xticks([])
        plt.tight_layout()
        img_dir = self.paths["img_output"]
        os.makedirs(img_dir, exist_ok=True)
        
        # Exclui o arquivo do dia anterior (day before) com o mesmo nome, mas data do day before (dbf_yesterday_str)
        dbf_yesterday = (self.yesterday_ts - pd.Timedelta(days=1)).date()
        dbf_yesterday_str = dbf_yesterday.strftime('%d-%m-%Y').replace("/", "-")
        dbf_image_path = os.path.join(img_dir, f'graficos_unidades_{dbf_yesterday_str}.png')
        if os.path.isfile(dbf_image_path):
            try:
                os.remove(dbf_image_path)
            except Exception as e:
                print(f'Erro ao excluir o arquivo de imagem do dia anterior: {e}')
        
        image_path = os.path.join(img_dir, f'graficos_unidades_{yesterday_str.replace("/", "-")}.png')
        plt.savefig(image_path, dpi=300, bbox_inches='tight')
        plt.close()
        print("[OK] Gráficos de unidades criados.")

    def automacao_envio_whatsapp(self):
        print("[INFO] Realizando automação para envio de imagens via WhatsApp (abrindo arquivos com pyautogui)...")
        img_dir = self.paths["img_output"]
        time.sleep(1)
        pyautogui.hotkey('win', 'e')
        time.sleep(4)
        pyautogui.hotkey('ctrl', 'l')
        time.sleep(1.5)
        pyautogui.typewrite(img_dir)
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(2.5)
        pyautogui.typewrite(r'graficos')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1.5)
        pyautogui.press('win')
        time.sleep(1.5)
        pyautogui.typewrite('whatsapp')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(10)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1.5)
        pyautogui.typewrite("raphael")
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(4)
        pyautogui.press('enter')
        time.sleep(1.5)
        # Segundo arquivo
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(2.5)
        pyautogui.typewrite(r'tabela')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1.5)
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(4)
        pyautogui.press('enter')
        print("[OK] Imagens enviadas pelo WhatsApp.\n")

    def salvar_relatorio(self):
        print("[INFO] Salvando relatório Excel no caminho padrão, na pasta de relatórios e no Sharepoint...")
        # Salva e faz cópia do relatório
        self.wb.save(self.paths["template_xlsx"])
        name_file = f'RELATORIO_ATIVACOES_CANCELAMENTOS_{self.yesterday}.xlsx'
        os.makedirs(self.paths["report_output"], exist_ok=True)
        output_full_path = os.path.join(self.paths["report_output"], name_file)
        shutil.copy(self.paths["template_xlsx"], output_full_path)

        name_file_2 = 'RELATORIO_ATIVACOES_CANCELAMENTOS.xlsx'
        full_path_sharepoint = os.path.join(self.paths["sharepoint_path"], name_file_2)
        shutil.copy(output_full_path, full_path_sharepoint)
        self.wb.close()
        print("[OK] Arquivo Excel salvo e copiado para pasta central e Sharepoint.")

    def executa_relatorio(self):
        print("\n==== INÍCIO EXECUÇÃO DO RELATÓRIO ====")
        self.update_base_sheet()
        self.preencher_relatorio()
        self.criar_imagem_resumo()
        self.criar_graficos_unidades()
        self.automacao_envio_whatsapp()
        self.salvar_relatorio()
        print("==== FIM EXECUÇÃO DO RELATÓRIO ====\n")

if __name__ == "__main__":
    print("Executando rotina de relatório diário...\n")
    relatorio = RelatorioAtivacoesCancelamentos()
    relatorio.executa_relatorio()
    print("Rotina finalizada.")