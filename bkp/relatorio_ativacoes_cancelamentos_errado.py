import os
import time
import datetime as dt
import pandas as pd
import awswrangler as awr
import openpyxl
import shutil
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import pyautogui

class RelatorioAtivacoesCancelamentos:
    PATHS = {
        "ativos": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_ATIVOS.sql",
        "cancelamentos_integrais": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_INTEGRAIS.sql",
        "cancelamentos_parciais": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_PARCIAIS.sql",
        "template_xlsx": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\template\TEMPLATE_BASE_ATIVACOES_CANCELAMENTOS.xlsx",
        "img": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\img",
        "reports": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\reports",
        "sharepoint": r"C:\Users\raphael.almeida\OneDrive - Grupo Unus\analise de dados - Arquivos em excel"
    }

    EMPRESAS = ["SEGTRUCK", "STCOOP", "VIAVANTE", "TAG"]

    def __init__(self):
        print("1) Iniciando rotina RelatorioAtivacoesCancelamentos")

        self.data_ref = self.get_yesterday()

        print("2) Lendo dados dos arquivos SQL (Ativos e Cancelamentos)...")
        self.df_ativos = self.get_dataframe_from_sql(self.PATHS["ativos"])
        self.df_cancelamentos_integrais = self.get_dataframe_from_sql(self.PATHS["cancelamentos_integrais"])
        self.df_cancelamentos_parciais = self.get_dataframe_from_sql(self.PATHS["cancelamentos_parciais"])

        print("3) Formatando e consolidando dados de cancelamentos...")
        self.format_all_types()
        self.concat_cancelamentos()

        print("4) Abrindo planilha Excel de template...")
        self.wb = openpyxl.load_workbook(self.PATHS["template_xlsx"])

        print("5) Preenchendo abas de ATIVOS e CANCELAMENTOS na planilha...")
        self.prepare_workbook_tabs()

        print("6) Calculando número de ativos e preenchendo aba BASE...")
        self.qtd_ativos = self.fill_base_tab()

        print("7) Calculando ativados/cancelados por empresa...")
        self.ativados, self.cancelados = self.get_ativados_cancelados_por_empresa()

        print("8) Preenchendo e salvando resumo na aba RELATORIO...")
        self.write_resume_tab()

        print("9) Gerando e salvando imagens/tabelas dos dados...")
        self.save_tabela_resumo_as_image()
        self.save_graficos_unidades()

        print("10) Salvando arquivo final e automatizando envio via WhatsApp...")
        self.save_and_export_workbook()
        self.automate_whatsapp_send()

    def get_yesterday(self):
        today = pd.Timestamp.today().normalize()
        return today - dt.timedelta(days=3) if today.weekday() == 0 else today - dt.timedelta(days=1)

    def normalize_date(self, d):
        if isinstance(d, (pd.Timestamp, dt.datetime)):
            return d.date()
        try:
            return pd.to_datetime(d).date()
        except Exception:
            return d

    def get_dataframe_from_sql(self, sql_path):
        with open(sql_path, 'r', encoding='utf-8') as f:
            sql = f.read()
        df = awr.athena.read_sql_query(sql=sql, database='silver')
        # drop column 'rn' if exists
        if "rn" in df.columns:
            df = df.drop(columns=['rn'])
        return df

    def format_type(self, df):
        for col in df.select_dtypes(include=["string", "object"]).columns:
            df[col] = df[col].astype(str).str.upper()
        return df

    def format_all_types(self):
        self.df_ativos = self.format_type(self.df_ativos)
        self.df_cancelamentos_integrais = self.format_type(self.df_cancelamentos_integrais)
        self.df_cancelamentos_parciais = self.format_type(self.df_cancelamentos_parciais)

    def concat_cancelamentos(self):
        self.df_cancelamentos_integrais['identificador'] = 'INTEGRAL'
        self.df_cancelamentos_parciais['identificador'] = 'PARCIAL'
        self.df_cancelamentos = pd.concat([
            self.df_cancelamentos_integrais, self.df_cancelamentos_parciais
        ], ignore_index=True)
        self.df_cancelamentos = self.df_cancelamentos.sort_values(
            by='data_cancelamento', ascending=False
        ).drop_duplicates(
            subset=['chassi', 'empresa', 'coverage_id'],
            keep='first'
        ).reset_index(drop=True)
        # Ensure dates are in correct format
        self.df_cancelamentos['data_cancelamento'] = pd.to_datetime(
            self.df_cancelamentos['data_cancelamento'], errors='coerce'
        ).dt.date
        self.df_ativos = self.df_ativos.sort_values(
            by=['inicio_vig', 'data_ativacao'], ascending=False
        ).reset_index(drop=True)

    def prepare_workbook_tabs(self):
        self.w1 = self.wb['ATIVOS']
        self.w2 = self.wb['CANCELAMENTOS']
        self.w3 = self.wb['BASE']
        self.w4 = self.wb['RELATORIO']
        self.clear_sheet(self.w1)
        self.clear_sheet(self.w2)
        self.write_dataframe_to_sheet(self.df_ativos, self.w1)
        self.write_dataframe_to_sheet(self.df_cancelamentos, self.w2)

    def clear_sheet(self, sheet):
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(1, max_row)

    def fillna_df(self, df):
        numeric_cols = df.select_dtypes(include=['number']).columns
        string_cols = df.select_dtypes(include=['object', 'string']).columns
        df[numeric_cols] = df[numeric_cols].fillna(0)
        df[string_cols] = df[string_cols].fillna('N/A')
        return df

    def write_dataframe_to_sheet(self, df, ws):
        df = self.fillna_df(df)
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        if not df.empty:
            for r_idx, row in enumerate(df.values, start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    def get_mask_period(self, col):
        today = dt.date.today()
        yesterday = today - dt.timedelta(days=1)
        if today.weekday() == 0:
            sexta = today - dt.timedelta(days=3)
            domingo = today - dt.timedelta(days=1)
            return (col >= sexta) & (col <= domingo)
        else:
            return col == yesterday

    def get_ativados_cancelados_por_empresa(self):
        ativados = {}
        cancelados = {}
        for e in self.EMPRESAS:
            ativos_mask = self.get_mask_period(self.df_ativos['inicio_vig']).astype(bool) & (self.df_ativos['empresa'] == e)
            ativados[e] = int(self.df_ativos[ativos_mask].shape[0])
            canc_mask = self.get_mask_period(self.df_cancelamentos['data_cancelamento']).astype(bool) & (self.df_cancelamentos['empresa'] == e)
            cancelados[e] = int(self.df_cancelamentos[canc_mask].shape[0])
        return ativados, cancelados

    def get_first_empty_row(self, ws, col=2):
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=col).value is None:
                return row
        return ws.max_row + 1

    def fill_base_tab(self):
        first_empty_row = self.get_first_empty_row(self.w3)
        data_atual_planilha = self.w3[f'A{first_empty_row}'].value

        # Normalize yesterday as date
        ontem = self.normalize_date(self.data_ref)
        if isinstance(data_atual_planilha, (pd.Timestamp, dt.datetime)):
            data_atual_planilha = data_atual_planilha.date()
        elif isinstance(data_atual_planilha, str):
            try:
                data_atual_planilha = pd.to_datetime(data_atual_planilha).date()
            except Exception:
                pass

        if data_atual_planilha == ontem:
            if "data_ativacao" in self.df_ativos.columns:
                ativ_col = pd.to_datetime(self.df_ativos["data_ativacao"], errors='coerce')
                df_ativos_filtrados = self.df_ativos[ativ_col <= pd.to_datetime(ontem)]
                qtd_ativos = df_ativos_filtrados['chassi'].nunique()
            else:
                qtd_ativos = self.df_ativos['chassi'].nunique()
            hoje = dt.date.today()
            if hoje.weekday() == 0 and first_empty_row >= 3:
                self.w3[f'B{first_empty_row - 2}'] = 'DOMINGO'
                self.w3[f'B{first_empty_row - 1}'] = 'SÁBADO'
                self.w3[f'B{first_empty_row}'] = qtd_ativos
            else:
                self.w3[f'B{first_empty_row}'] = qtd_ativos
            return qtd_ativos
        else:
            # fallback caso não haja preenchimento
            return self.df_ativos['chassi'].nunique()

    def write_resume_tab(self):
        hoje = dt.date.today()
        yest_date = self.normalize_date(self.data_ref)
        if hoje.weekday() == 0:
            sexta = yest_date - dt.timedelta(days=2)
            domingo = yest_date
            self.w4['C2'] = f"{sexta.strftime('%d/%m/%Y')} (sexta) - {domingo.strftime('%d/%m/%Y')} (domingo)"
        else:
            self.w4['C2'] = yest_date.strftime('%d/%m/%Y')
        self.w4['C3'] = self.qtd_ativos
        # Ativados
        for idx, e in enumerate(self.EMPRESAS, start=6):
            self.w4[f'C{idx}'] = self.ativados.get(e, 0)
            self.w4[f'D{idx}'] = self.cancelados.get(e, 0)

    def save_tabela_resumo_as_image(self):
        ativados = [self.ativados[e] for e in self.EMPRESAS]
        cancelados = [self.cancelados[e] for e in self.EMPRESAS]
        total_ativados = sum(ativados)
        total_cancelados = sum(cancelados)
        total_ativos_geral = f"{self.qtd_ativos:,}".replace(',', '.')

        row_dados_main = [[e.capitalize(), a, c] for e, a, c in zip(self.EMPRESAS, ativados, cancelados)]
        row_dados_main.append(['Total', total_ativados, total_cancelados])
        data_atual = self.normalize_date(self.data_ref)
        data_atual_str = data_atual.strftime('%d/%m/%Y')
        dados_topo = [ ['DATA', data_atual_str], ['TOTAL ATIVOS', total_ativos_geral] ]
        colunas_main = ['EMPRESA', 'ATIVADOS', 'CANCELADOS']

        fig, ax = plt.subplots(figsize=(6, 5))
        ax.axis('off')
        table_top = ax.table(
            cellText=dados_topo,
            cellLoc='center',
            loc='center',
            bbox=[0.0, 0.75, 1.0, 0.2]
        )
        cor_fundo_cinza = '#E6E6E6'
        cor_fundo_branco = '#FFFFFF'

        for (row, col), cell in table_top.get_celld().items():
            cell.set_height(0.1)
            cell.set_text_props(fontweight='bold', color='#000000')
            cell.set_facecolor(cor_fundo_cinza if col == 0 else cor_fundo_branco)

        table_main = ax.table(
            cellText=row_dados_main,
            colLabels=colunas_main,
            cellLoc='center',
            loc='center',
            bbox=[0.0, 0.0, 1.0, 0.7]
        )
        for (row, col), cell in table_main.get_celld().items():
            cell.set_height(0.12)
            if row == 0:
                cell.set_facecolor(cor_fundo_cinza)
                cell.set_text_props(fontweight='bold', color='#000000')
            elif col == 0:
                cell.set_facecolor(cor_fundo_cinza)
                cell.set_text_props(fontweight='bold', color='#000000')
            elif row == 3:
                cell.set_facecolor(cor_fundo_branco)
                cell.set_text_props(fontweight='normal', color='#000000')
            elif row == 5:
                cell.set_facecolor(cor_fundo_cinza)
                cell.set_text_props(fontweight='bold', color='#000000')
            else:
                cell.set_facecolor(cor_fundo_branco)
                cell.set_text_props(fontweight='normal', color='#000000')

        plt.tight_layout()
        data_antiga = (data_atual - dt.timedelta(days=1)).strftime('%d/%m/%Y')
        image_path_antiga = os.path.join(
            self.PATHS["img"], f'tabela_ativacoes_cancelamentos_{data_antiga.replace("/", "-")}.png'
        )
        if os.path.exists(image_path_antiga):
            try:
                os.remove(image_path_antiga)
            except Exception as e:
                print(f"Erro ao remover arquivo antigo: {image_path_antiga}. Erro: {e}")

        os.makedirs(self.PATHS["img"], exist_ok=True)
        image_path = os.path.join(self.PATHS["img"], f'tabela_ativacoes_cancelamentos_{data_atual_str.replace("/", "-")}.png')
        plt.savefig(image_path, dpi=300, bbox_inches='tight')
        plt.close()

    def save_graficos_unidades(self):
        yest_date = self.normalize_date(self.data_ref)

        df_ativos_ontem = self.df_ativos[
            (pd.to_datetime(self.df_ativos['inicio_vig'], errors='coerce').dt.date == yest_date)
        ]
        df_ativos_ontem_unidades = df_ativos_ontem.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)
        df_cancelamentos_ontem = self.df_cancelamentos[
            (self.df_cancelamentos['data_cancelamento'] == yest_date)
        ]
        df_cancelamentos_ontem_unidades = df_cancelamentos_ontem.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)

        data_str = yest_date.strftime('%d/%m/%Y')

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 6))
        # Ativos
        if not df_ativos_ontem_unidades.empty:
            df_ativos_ontem_unidades.plot(kind='barh', ax=ax1, color='#2ecc71', edgecolor='black', linewidth=0.5)
            ax1.set_title(f'Ativos por Unidade - {data_str}', fontsize=14, fontweight='bold', pad=15)
            ax1.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax1.set_ylabel('', fontsize=11, fontweight='bold')
            ax1.set_xticks([])
            ax1.grid(axis='x', alpha=0.3, linestyle='--', visible=False)
            ax1.invert_yaxis()
            xmax = df_ativos_ontem_unidades.values.max() if len(df_ativos_ontem_unidades) > 0 else 1
            ax1.set_xlim(0, xmax + max(3, int(0.15 * xmax)))
            for i, v in enumerate(df_ativos_ontem_unidades.values):
                ax1.text(v + 0.7, i, str(int(v)), va='center', fontweight='bold')
        else:
            ax1.text(0.5, 0.5, 'Sem dados disponíveis', ha='center', va='center', transform=ax1.transAxes, fontsize=12)
            ax1.set_title(f'Ativos por Unidade - {data_str}', fontsize=14, fontweight='bold')
            ax1.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax1.set_ylabel('', fontsize=11, fontweight='bold')
            ax1.set_xticks([])

        # Cancelados
        if not df_cancelamentos_ontem_unidades.empty:
            df_cancelamentos_ontem_unidades.plot(kind='barh', ax=ax2, color='#e74c3c', edgecolor='black', linewidth=0.5)
            ax2.set_title(f'Cancelados por Unidade - {data_str}', fontsize=14, fontweight='bold', pad=15)
            ax2.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax2.set_ylabel('', fontsize=11, fontweight='bold')
            ax2.set_xticks([])
            ax2.grid(axis='x', alpha=0.3, linestyle='--', visible=False)
            ax2.invert_yaxis()
            xmax2 = df_cancelamentos_ontem_unidades.values.max() if len(df_cancelamentos_ontem_unidades) > 0 else 1
            ax2.set_xlim(0, xmax2 + max(3, int(0.15 * xmax2)))
            for i, v in enumerate(df_cancelamentos_ontem_unidades.values):
                ax2.text(v + 0.7, i, str(int(v)), va='center', fontweight='bold')
        else:
            ax2.text(0.5, 0.5, 'Sem dados disponíveis', ha='center', va='center', transform=ax2.transAxes, fontsize=12)
            ax2.set_title(f'Cancelados por Unidade - {data_str}', fontsize=14, fontweight='bold')
            ax2.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax2.set_ylabel('', fontsize=11, fontweight='bold')
            ax2.set_xticks([])
        plt.tight_layout()
        os.makedirs(self.PATHS["img"], exist_ok=True)
        image_path = os.path.join(self.PATHS["img"], f'graficos_unidades_{data_str.replace("/", "-")}.png')
        # Antes de salvar, exclui arquivo antigo se já existir
        if os.path.isfile(image_path):
            try:
                os.remove(image_path)
            except Exception as e:
                print(f'Erro ao excluir o arquivo de imagem existente: {e}')
        # Exclui o arquivo do dia anterior (day before) com o mesmo nome, mas data do day before (dbf_yesterday_str)
        dbf_yesterday = (self.data_ref - pd.Timedelta(days=1)).date()
        dbf_yesterday_str = dbf_yesterday.strftime('%d-%m-%Y').replace("/", "-")
        dbf_image_path = os.path.join(self.PATHS["img"], f'graficos_unidades_{dbf_yesterday_str}.png')
        if os.path.isfile(dbf_image_path):
            try:
                os.remove(dbf_image_path)
            except Exception as e:
                print(f'Erro ao excluir o arquivo de imagem do dia anterior: {e}')
        plt.savefig(image_path, dpi=300, bbox_inches='tight')
        plt.close()

    def automate_whatsapp_send(self):
        """Automatiza abrir pasta de imagens, copiar gráficos e enviar pelo whatsapp. Mantém à parte: 
        AVISO: Uso de pyautogui é frágil fora de ambiente controlado."""
        # Abrir pasta de imagens
        time.sleep(1)
        pyautogui.hotkey('win', 'e')
        time.sleep(4)
        pyautogui.hotkey('ctrl', 'l')
        time.sleep(1.5)
        pyautogui.typewrite(self.PATHS["img"])
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(2.5)
        pyautogui.typewrite('graficos')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        #whatsapp - primeiro arquivo
        time.sleep(1.5)
        pyautogui.press('win')
        time.sleep(1.5)
        pyautogui.typewrite('whatsapp')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1.5)
        pyautogui.typewrite("raphael")
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        #whatsapp - segundo arquivo
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(2.5)
        pyautogui.typewrite('tabela')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1.5)
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2.5)
        pyautogui.press('enter')

    def save_and_export_workbook(self):
        self.wb.save(self.PATHS["template_xlsx"])
        data_str = self.data_ref.strftime('%d-%m-%Y') if hasattr(self.data_ref, 'strftime') else str(self.data_ref)
        name_file = f'RELATORIO_ATIVACOES_CANCELAMENTOS_{data_str}.xlsx'
        os.makedirs(self.PATHS["reports"], exist_ok=True)
        output_full_path = os.path.join(self.PATHS["reports"], name_file)
        shutil.copy(self.PATHS["template_xlsx"], output_full_path)
        name_file_2 = 'RELATORIO_ATIVACOES_CANCELAMENTOS.xlsx'
        full_path_sharepoint = os.path.join(self.PATHS["sharepoint"], name_file_2)
        shutil.copy(output_full_path, full_path_sharepoint)
        self.wb.close()

if __name__ == "__main__":
    RelatorioAtivacoesCancelamentos()