import os
import time
import datetime as dt
import pandas as pd
import awswrangler as awr
import openpyxl
import shutil
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import pyautogui

class ETL_relatorio_ativ_cancel:
    PATHS = {
        "ativos": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_ATIVOS.sql",
        "cancelamentos_integrais": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_INTEGRAIS.sql",
        "cancelamentos_parciais": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_PARCIAIS.sql",
        "template_xlsx": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\template\TEMPLATE_BASE_ATIVACOES_CANCELAMENTOS.xlsx",
        "img": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\img",
        "reports": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\reports",
        "sharepoint": r"C:\Users\raphael.almeida\OneDrive - Grupo Unus\analise de dados - Arquivos em excel"
    }

    EMPRESAS = ["SEGTRUCK", "STCOOP", "VIAVANTE", "TAG"]

    def __init__(self):
        self.df_ativos = None
        self.df_cancelamentos_integrais = None
        self.df_cancelamentos_parciais = None
        self.df_cancelamentos = None
        self.yesterday = self.get_yesterday()
        self.hoje = dt.date.today()
        self.qtd_ativos = 0
        self.ativados = {}
        self.cancelados = {}

    def run(self):
        self.carregar_dados()
        self.processar_dados()
        self.preencher_planilha()
        self.criar_imagem_resumo()
        self.criar_imagem_unidades()
        self.automatizar_whatsapp()
        self.salvar_e_compartilhar()

    @staticmethod
    def format_type(df):
        # Excluir colunas de data da conversão para maiúsculo
        date_columns = ['inicio_vig', 'data_ativacao', 'data_cancelamento', 'data_registro', 'data_extracao']
        for col in df.select_dtypes(include=["string", "object"]).columns:
            # Pular colunas que são datas mas estão como object/string temporariamente
            if col not in date_columns:
                df[col] = df[col].astype(str).str.upper()
        return df

    def get_yesterday(self):
        today = pd.Timestamp.today().normalize()
        return (today - dt.timedelta(days=3) if today.weekday() == 0 else today - dt.timedelta(days=1)).date()

    def carregar_dados(self):
        with open(self.PATHS["ativos"], 'r', encoding='utf-8') as f:
            sql_ativos = f.read()
        self.df_ativos = awr.athena.read_sql_query(sql=sql_ativos, database='silver')
        if 'rn' in self.df_ativos.columns:
            self.df_ativos = self.df_ativos.drop(columns=['rn'])
        self.df_ativos = (self.df_ativos
            .sort_values(by=['chassi', 'inicio_vig', 'data_ativacao'], ascending=[True, False, False])
            .drop_duplicates(subset=['chassi'], keep='first')
        )
        with open(self.PATHS["cancelamentos_integrais"], 'r', encoding='utf-8') as f:
            sql_cancelamentos_integrais = f.read()
        self.df_cancelamentos_integrais = awr.athena.read_sql_query(sql=sql_cancelamentos_integrais, database='silver')

        with open(self.PATHS["cancelamentos_parciais"], 'r', encoding='utf-8') as f:
            sql_cancelamentos_parciais = f.read()
        self.df_cancelamentos_parciais = awr.athena.read_sql_query(sql=sql_cancelamentos_parciais, database='silver')

        # Padroniza tipos de string para maiúsculo
        self.df_ativos = self.format_type(self.df_ativos)
        self.df_cancelamentos_integrais = self.format_type(self.df_cancelamentos_integrais)
        self.df_cancelamentos_parciais = self.format_type(self.df_cancelamentos_parciais)

    def processar_dados(self):
        today = dt.date.today()

        # Converter inicio_vig para date se necessário
        if 'inicio_vig' in self.df_ativos.columns:
            print(f"[DEBUG] Tipo de inicio_vig antes da conversão: {self.df_ativos['inicio_vig'].dtype}")
            self.df_ativos['inicio_vig'] = pd.to_datetime(
                self.df_ativos['inicio_vig'], errors='coerce'
            ).dt.date
            print(f"[DEBUG] Tipo de inicio_vig depois da conversão: {type(self.df_ativos['inicio_vig'].iloc[0]) if len(self.df_ativos) > 0 else 'vazio'}")
            print(f"[DEBUG] Exemplo de valores de inicio_vig: {self.df_ativos['inicio_vig'].head(3).tolist() if len(self.df_ativos) > 0 else 'vazio'}")

        self.df_cancelamentos_integrais['data_cancelamento'] = pd.to_datetime(
            self.df_cancelamentos_integrais['data_cancelamento'], errors='coerce'
        ).dt.date

        self.df_cancelamentos_integrais['identificador'] = 'INTEGRAL'
        self.df_cancelamentos_parciais['identificador'] = 'PARCIAL'
        self.df_cancelamentos = pd.concat(
            [self.df_cancelamentos_integrais, self.df_cancelamentos_parciais], ignore_index=True
        )
        self.df_cancelamentos = (
            self.df_cancelamentos
            .sort_values(by='data_cancelamento', ascending=False)
            .drop_duplicates(subset=['chassi', 'empresa', 'coverage_id'], keep='first')
            .reset_index(drop=True)
        )
        self.df_ativos = (self.df_ativos
            .sort_values(by=['inicio_vig', 'data_ativacao'], ascending=False)
            .reset_index(drop=True)
        )

        if today.weekday() == 0:
            sexta = today - dt.timedelta(days=3)
            domingo = today - dt.timedelta(days=1)
            print(f"[DEBUG] Segunda-feira: filtrando entre {sexta} e {domingo}")
            self.ativos_mask = (self.df_ativos['inicio_vig'] >= sexta) & (self.df_ativos['inicio_vig'] <= domingo)
            self.cancelamentos_mask = (self.df_cancelamentos['data_cancelamento'] >= sexta) & (self.df_cancelamentos['data_cancelamento'] <= domingo)
        else:
            yesterday = today - dt.timedelta(days=1)
            print(f"[DEBUG] Dia comum: filtrando para {yesterday}")
            print(f"[DEBUG] Tipo de yesterday: {type(yesterday)}")
            self.ativos_mask = (self.df_ativos['inicio_vig'] == yesterday)
            self.cancelamentos_mask = (self.df_cancelamentos['data_cancelamento'] == yesterday)

        print(f"[DEBUG] Total de registros que passaram na máscara de ativos: {self.ativos_mask.sum()}")
        print(f"[DEBUG] Empresas únicas nos ativos filtrados: {self.df_ativos[self.ativos_mask]['empresa'].value_counts().to_dict() if self.ativos_mask.sum() > 0 else 'nenhum'}")

        self.ativados = {
            'SEGTRUCK': len(self.df_ativos[(self.df_ativos['empresa'] == 'SEGTRUCK') & self.ativos_mask]),
            'STCOOP': len(self.df_ativos[(self.df_ativos['empresa'] == 'STCOOP') & self.ativos_mask]),
            'VIAVANTE': len(self.df_ativos[(self.df_ativos['empresa'] == 'VIAVANTE') & self.ativos_mask]),
            'TAG': len(self.df_ativos[(self.df_ativos['empresa'] == 'TAG') & self.ativos_mask])
        }
        self.cancelados = {
            'SEGTRUCK': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'SEGTRUCK') & self.cancelamentos_mask]),
            'STCOOP': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'STCOOP') & self.cancelamentos_mask]),
            'VIAVANTE': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'VIAVANTE') & self.cancelamentos_mask]),
            'TAG': len(self.df_cancelamentos[(self.df_cancelamentos['empresa'] == 'TAG') & self.cancelamentos_mask])
        }
        
        print(f"[DEBUG] Resultado final de ativados: {self.ativados}")
        print(f"[DEBUG] Resultado final de cancelados: {self.cancelados}")

    def clear_sheet(self, sheet):
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(1, max_row)

    def preencher_planilha(self):
        template = self.PATHS["template_xlsx"]
        wb = openpyxl.load_workbook(template)

        w1 = wb['ATIVOS']
        w2 = wb['CANCELAMENTOS']
        w3 = wb['BASE']
        w4 = wb['RELATORIO']

        self.clear_sheet(w1)
        self.clear_sheet(w2)

        for df, ws in [(self.df_ativos, w1), (self.df_cancelamentos, w2)]:
            numeric_cols = df.select_dtypes(include=['number']).columns
            string_cols = df.select_dtypes(include=['object', 'string']).columns
            df[numeric_cols] = df[numeric_cols].fillna(0)
            df[string_cols] = df[string_cols].fillna('N/A')

        for c_idx, col_name in enumerate(self.df_ativos.columns, start=1):
            w1.cell(row=1, column=c_idx, value=col_name)
        if not self.df_ativos.empty:
            for r_idx, row in enumerate(self.df_ativos.values, start=2):
                for c_idx, value in enumerate(row, start=1):
                    w1.cell(row=r_idx, column=c_idx, value=value)
        for c_idx, col_name in enumerate(self.df_cancelamentos.columns, start=1):
            w2.cell(row=1, column=c_idx, value=col_name)
        if not self.df_cancelamentos.empty:
            for r_idx, row in enumerate(self.df_cancelamentos.values, start=2):
                for c_idx, value in enumerate(row, start=1):
                    w2.cell(row=r_idx, column=c_idx, value=value)

        # Encontrar primeira linha vazia na coluna B da planilha BASE
        first_empty_row = 1
        for row in range(1, w3.max_row + 1):
            if w3.cell(row=row, column=2).value is None:
                first_empty_row = row
                break
        else:
            first_empty_row = w3.max_row + 1

        data_atual_planilha = w3['A' + str(first_empty_row)].value

        if isinstance(data_atual_planilha, pd.Timestamp):
            data_atual_planilha = data_atual_planilha.date()
        elif hasattr(data_atual_planilha, 'date'):
            data_atual_planilha = data_atual_planilha.date()
        elif isinstance(data_atual_planilha, dt.datetime):
            data_atual_planilha = data_atual_planilha.date()
        elif isinstance(data_atual_planilha, str):
            try:
                data_atual_planilha = pd.to_datetime(data_atual_planilha).date()
            except:
                data_atual_planilha = data_atual_planilha

        yesterday = self.yesterday

        yesterday_date = yesterday
        if isinstance(yesterday, pd.Timestamp):
            yesterday_date = yesterday.date()
        elif isinstance(yesterday, dt.datetime):
            yesterday_date = yesterday.date()

        if data_atual_planilha == yesterday_date:
            if "data_ativacao" in self.df_ativos.columns:
                data_ativacao_col = pd.to_datetime(self.df_ativos["data_ativacao"], errors='coerce')
                df_ativos_filtrados = self.df_ativos[data_ativacao_col <= pd.to_datetime(yesterday_date)]
                qtd_ativos = df_ativos_filtrados['chassi'].nunique()
            else:
                qtd_ativos = self.df_ativos['chassi'].nunique()
            if self.hoje.weekday() == 0 and first_empty_row >= 3:
                w3['B' + str(first_empty_row - 2)] = 'DOMINGO'
                w3['B' + str(first_empty_row - 1)] = 'SÁBADO'
                w3['B' + str(first_empty_row)] = qtd_ativos
            else:
                w3['B' + str(first_empty_row)] = qtd_ativos
            self.qtd_ativos = qtd_ativos

        else:
            self.qtd_ativos = self.df_ativos['chassi'].nunique()

        # Resumo
        hoje = dt.date.today()
        dia_semana = hoje.weekday()
        yest_date = yesterday_date
        w4['C3'] = self.qtd_ativos

        if dia_semana == 0:
            sexta = yest_date - dt.timedelta(days=2)
            domingo = yest_date
            sexta_str = sexta.strftime('%d/%m/%Y')
            domingo_str = domingo.strftime('%d/%m/%Y')
            resumo_periodo = f"{sexta_str} (sexta) - {domingo_str} (domingo)"
            w4['C2'] = resumo_periodo
        else:
            w4['C2'] = yest_date.strftime('%d/%m/%Y')

        w4['C6'] = self.ativados['SEGTRUCK']
        w4['C7'] = self.ativados['STCOOP']
        w4['C8'] = self.ativados['VIAVANTE']
        w4['C9'] = self.ativados['TAG']

        w4['D6'] = self.cancelados['SEGTRUCK']
        w4['D7'] = self.cancelados['STCOOP']
        w4['D8'] = self.cancelados['VIAVANTE']
        w4['D9'] = self.cancelados['TAG']

        self.wb = wb  # Para salvar depois
        self.w4 = w4

    def criar_imagem_resumo(self):
        df_ativos = self.df_ativos
        qtd_ativos = self.qtd_ativos
        ativados = self.ativados
        cancelados = self.cancelados
        yesterday = self.yesterday
        hoje = dt.date.today()
        dia_semana = hoje.weekday()
        yest_date = yesterday
        total_ativados = sum(ativados.values())
        total_cancelados = sum(cancelados.values())

        if dia_semana == 0:
            sexta = yest_date - dt.timedelta(days=2)
            domingo = yest_date
            sexta_str = sexta.strftime('%d/%m/%Y')
            domingo_str = domingo.strftime('%d/%m/%Y')
            data_atual = f"{sexta_str} (sexta) - {domingo_str} (domingo)"
        else:
            data_atual = yest_date.strftime('%d/%m/%Y')

        total_ativos_geral = f"{qtd_ativos:,}".replace(',', '.')

        # Estruturação das tabelas
        cor_fundo_cinza = '#E6E6E6'
        cor_fundo_branco = '#FFFFFF'

        colunas_main = ['EMPRESA', 'ATIVADOS', 'CANCELADOS']
        dados_main = [
            ['Segtruck', ativados['SEGTRUCK'], cancelados['SEGTRUCK']],
            ['Stcoop', ativados['STCOOP'], cancelados['STCOOP']],
            ['Viavante', ativados['VIAVANTE'], cancelados['VIAVANTE']],
            ['Tag', ativados['TAG'], cancelados['TAG']],
            ['Total', total_ativados, total_cancelados]
        ]

        dados_topo = [
            ['DATA', data_atual],
            ['TOTAL ATIVOS', total_ativos_geral]
        ]

        fig, ax = plt.subplots(figsize=(6, 5))
        ax.axis('off')

        table_top = ax.table(
            cellText=dados_topo,
            cellLoc='center',
            loc='center',
            bbox=[0.0, 0.75, 1.0, 0.2]
        )
        for (row, col), cell in table_top.get_celld().items():
            cell.set_height(0.1)
            cell.set_text_props(fontweight='bold', color='#000000')
            cell.set_facecolor(cor_fundo_cinza if col == 0 else cor_fundo_branco)

        table_main = ax.table(
            cellText=dados_main,
            colLabels=colunas_main,
            cellLoc='center',
            loc='center',
            bbox=[0.0, 0.0, 1.0, 0.7]
        )

        for (row, col), cell in table_main.get_celld().items():
            cell.set_height(0.12)
            if row == 0:
                cell.set_facecolor(cor_fundo_cinza)
                cell.set_text_props(fontweight='bold', color='#000000')
            elif col == 0:
                cell.set_facecolor(cor_fundo_cinza)
                cell.set_text_props(fontweight='bold', color='#000000')
            elif row == 3:
                cell.set_facecolor(cor_fundo_branco)
                cell.set_text_props(fontweight='normal', color='#000000')
            elif row == 5:
                cell.set_facecolor(cor_fundo_cinza)
                cell.set_text_props(fontweight='bold', color='#000000')
            else:
                cell.set_facecolor(cor_fundo_branco)
                cell.set_text_props(fontweight='normal', color='#000000')

        plt.tight_layout()
        os.makedirs(self.PATHS["img"], exist_ok=True)

        # Remove previous day's image before saving today's
        try:
            data_antiga = (dt.datetime.strptime(data_atual, "%d/%m/%Y") - dt.timedelta(days=1)).strftime("%d/%m/%Y")
            image_path_antiga = os.path.join(
                self.PATHS["img"], f'tabela_ativacoes_cancelamentos_{data_antiga.replace("/", "-")}.png'
            )
            if os.path.exists(image_path_antiga):
                os.remove(image_path_antiga)
        except Exception as e:
            print(f"Erro ao remover arquivo antigo: {e}")

        image_path = os.path.join(
            self.PATHS["img"], f'tabela_ativacoes_cancelamentos_{data_atual.replace("/", "-")}.png'
        )
        plt.savefig(image_path, dpi=300, bbox_inches='tight')
        print(f'Tabela salva em: {image_path}')
        plt.close()

    def criar_imagem_unidades(self):
        print("[1/10] Iniciando geração dos gráficos de unidades (ativos e cancelados)...")
        df_ativos = self.df_ativos
        df_cancelamentos = self.df_cancelamentos
        yesterday = self.yesterday
        hoje = dt.date.today()
        dia_semana = hoje.weekday()
        print(f"[2/10] Referência de data: hoje={hoje.strftime('%d/%m/%Y')} (weekday={dia_semana}), referência processamento={yesterday.strftime('%d/%m/%Y')}")

        if dia_semana == 0:
            sexta = yesterday - dt.timedelta(days=2)
            print(f"[3/10] Hoje é segunda-feira, incluindo dados de sexta ({sexta.strftime('%d/%m/%Y')}), sábado e domingo.")
            period_mask = df_ativos["inicio_vig"].isin([sexta, sexta + dt.timedelta(days=1), yesterday])
            df_ativos_periodo = df_ativos[period_mask]
            df_ativos_ontem_unidades = df_ativos_periodo.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)

            period_mask_cancel = df_cancelamentos["data_cancelamento"].isin([sexta, sexta + dt.timedelta(days=1), yesterday])
            df_cancelamentos_periodo = df_cancelamentos[period_mask_cancel]
            df_cancelamentos_ontem_unidades = df_cancelamentos_periodo.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)

            sexta_str = sexta.strftime('%d/%m/%Y')
            domingo_str = yesterday.strftime('%d/%m/%Y')
            yesterday_str = f"{sexta_str} (sexta) - {domingo_str} (domingo)"
            print(f"[4/10] Período de referência para gráficos: {yesterday_str}")
        else:
            print(f"[3/10] Dia comum (não é segunda). Considerando apenas o dia {yesterday.strftime('%d/%m/%Y')}")
            df_ativos_ontem = df_ativos[df_ativos['inicio_vig'] == yesterday]
            df_ativos_ontem_unidades = df_ativos_ontem.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)
            df_cancelamentos_ontem = df_cancelamentos[df_cancelamentos['data_cancelamento'] == yesterday]
            df_cancelamentos_ontem_unidades = df_cancelamentos_ontem.groupby('unidade')['chassi'].nunique().sort_values(ascending=False)
            yesterday_str = yesterday.strftime('%d/%m/%Y')
            print(f"[4/10] Período de referência para gráficos: {yesterday_str}")

        print(f"[5/10] Montando gráfico. Quantidade de unidades com ativos: {len(df_ativos_ontem_unidades)}")
        print(f"[6/10] Montando gráfico. Quantidade de unidades com cancelamentos: {len(df_cancelamentos_ontem_unidades)}")

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 6))

        # Ativos
        if not df_ativos_ontem_unidades.empty:
            print("[7/10] Gerando gráfico de ATIVOS por unidade...")
            df_ativos_ontem_unidades.plot(kind='barh', ax=ax1, color='#2ecc71', edgecolor='black', linewidth=0.5)
            ax1.set_title(f'Ativos por Unidade - {yesterday_str}', fontsize=14, fontweight='bold', pad=15)
            ax1.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax1.set_ylabel('', fontsize=11, fontweight='bold')
            ax1.set_xticks([])
            ax1.grid(axis='x', alpha=0.3, linestyle='--', visible=False)
            ax1.invert_yaxis()
            xmax = df_ativos_ontem_unidades.values.max() if len(df_ativos_ontem_unidades) > 0 else 1
            ax1.set_xlim(0, xmax + max(3, int(0.15 * xmax)))
            for i, v in enumerate(df_ativos_ontem_unidades.values):
                ax1.text(v + 0.7, i, str(int(v)), va='center', fontweight='bold')
        else:
            print("[7/10] Não há dados de ATIVOS para o período, exibindo mensagem no gráfico.")
            ax1.text(0.5, 0.5, 'Sem dados disponíveis', ha='center', va='center', transform=ax1.transAxes, fontsize=12)
            ax1.set_title(f'Ativos por Unidade - {yesterday_str}', fontsize=14, fontweight='bold')
            ax1.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax1.set_ylabel('', fontsize=11, fontweight='bold')
            ax1.set_xticks([])

        # Cancelados
        if not df_cancelamentos_ontem_unidades.empty:
            print("[8/10] Gerando gráfico de CANCELADOS por unidade...")
            df_cancelamentos_ontem_unidades.plot(kind='barh', ax=ax2, color='#e74c3c', edgecolor='black', linewidth=0.5)
            ax2.set_title(f'Cancelados por Unidade - {yesterday_str}', fontsize=14, fontweight='bold', pad=15)
            ax2.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax2.set_ylabel('', fontsize=11, fontweight='bold')
            ax2.set_xticks([])
            ax2.grid(axis='x', alpha=0.3, linestyle='--', visible=False)
            ax2.invert_yaxis()
            xmax2 = df_cancelamentos_ontem_unidades.values.max() if len(df_cancelamentos_ontem_unidades) > 0 else 1
            ax2.set_xlim(0, xmax2 + max(3, int(0.15 * xmax2)))
            for i, v in enumerate(df_cancelamentos_ontem_unidades.values):
                ax2.text(v + 0.7, i, str(int(v)), va='center', fontweight='bold')
        else:
            print("[8/10] Não há dados de CANCELAMENTOS para o período, exibindo mensagem no gráfico.")
            ax2.text(0.5, 0.5, 'Sem dados disponíveis', ha='center', va='center', transform=ax2.transAxes, fontsize=12)
            ax2.set_title(f'Cancelados por Unidade - {yesterday_str}', fontsize=14, fontweight='bold')
            ax2.set_xlabel('Quantidade de Chassis', fontsize=11, fontweight='bold')
            ax2.set_ylabel('', fontsize=11, fontweight='bold')
            ax2.set_xticks([])

        plt.tight_layout()
        os.makedirs(self.PATHS["img"], exist_ok=True)
        image_path = os.path.join(self.PATHS["img"], f'graficos_unidades_{yesterday_str.replace("/", "-")}.png')
        print(f"[9/10] Caminho para salvar a imagem dos gráficos: {image_path}")

        # Exclui imagem do dia anterior, se existir
        if "(" in yesterday_str:
            # Caso sexta-domingo, extrai apenas a última data (domingo) para basear o dia anterior
            try:
                import re
                ultimo_data = re.findall(r'\((.*?)\)', yesterday_str)
                if ultimo_data and len(ultimo_data) >= 1:
                    data_ref = ultimo_data[-1].split()[0]
                    data_ref = dt.datetime.strptime(data_ref, "%d/%m/%Y").date()
                else:
                    data_ref = dt.datetime.strptime(yesterday_str[-10:], "%d/%m/%Y").date()
                print(f"[10/10] Período especial ('sexta-domingo'). Última data usada para imagem antiga: {data_ref.strftime('%d/%m/%Y')}")
            except Exception as e:
                data_ref = dt.date.today()
                print(f"[10/10] Falha ao extrair data para remover imagem anterior (sexta-domingo): {e}")
        else:
            try:
                data_ref = dt.datetime.strptime(yesterday_str, "%d/%m/%Y").date()
                print(f"[10/10] Usando data anterior padrão para remoção da imagem antiga: {data_ref.strftime('%d/%m/%Y')}")
            except Exception as e:
                data_ref = dt.date.today()
                print(f"[10/10] Falha ao extrair data para remover imagem anterior: {e}")

        dia_anterior = data_ref - dt.timedelta(days=1)
        dia_anterior_str = dia_anterior.strftime('%d/%m/%Y')
        image_path_old = os.path.join(self.PATHS["img"], f'graficos_unidades_{dia_anterior_str.replace("/", "-")}.png')
        if os.path.exists(image_path_old):
            try:
                os.remove(image_path_old)
                print(f"[LOG] Imagem antiga removida: {image_path_old}")
            except Exception as e:
                print(f"[LOG] Erro ao remover imagem antiga: {image_path_old}. Erro: {e}")
        else:
            print(f"[LOG] Nenhuma imagem antiga encontrada para remover em {image_path_old}")

        plt.savefig(image_path, dpi=300, bbox_inches='tight')
        print(f"[LOG] Gráficos de unidades salvos em: {image_path}")
        plt.close()
        print("[LOG] Finalizada a rotina dos gráficos de unidades.")

    def automatizar_whatsapp(self):
        time.sleep(1)
        pyautogui.hotkey('win', 'e')
        time.sleep(4)
        pyautogui.hotkey('ctrl', 'l')
        time.sleep(1.5)
        pyautogui.typewrite(self.PATHS["img"])
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(2.5)
        pyautogui.typewrite(r'graficos')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        #whatsapp - primeiro arquivo
        time.sleep(1.5)
        pyautogui.press('win')
        time.sleep(1.5)
        pyautogui.typewrite('whatsapp')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1.5)
        pyautogui.typewrite("# RELAT")
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        #whatsapp - segundo arquivo
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(2.5)
        pyautogui.typewrite(r'tabela')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1.5)
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2.5)
        pyautogui.press('enter')

    def salvar_e_compartilhar(self):
        self.wb.save(self.PATHS["template_xlsx"])
        name_file = f'RELATORIO_ATIVACOES_CANCELAMENTOS_{self.yesterday}.xlsx'
        output_path = self.PATHS["reports"]
        os.makedirs(output_path, exist_ok=True)
        output_full_path = os.path.join(output_path, name_file)
        shutil.copy(self.PATHS["template_xlsx"], output_full_path)
        name_file_2 = 'RELATORIO_ATIVACOES_CANCELAMENTOS.xlsx'
        full_path_sharepoint = os.path.join(self.PATHS["sharepoint"], name_file_2)
        shutil.copy(output_full_path, full_path_sharepoint)
        self.wb.close()

    @classmethod
    def run_load(cls):
        relatorio = cls()
        relatorio.run()

if __name__ == "__main__":
    ETL_relatorio_ativ_cancel.run_load()