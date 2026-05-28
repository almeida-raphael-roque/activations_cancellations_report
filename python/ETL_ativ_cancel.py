import pandas as pd
import datetime as dt
import awswrangler as awr
import openpyxl
import os
import shutil
import pyautogui
import time
from html2image import Html2Image

class ETL_relat_ativ_cancel:
    PATHS = {
        "ativos": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_ATIVOS.sql",
        "cancelamentos_integrais": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_INTEGRAIS.sql",
        "cancelamentos_parciais": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\sql\all_boards_CANCELAMENTOS_PARCIAIS.sql",
        "faturas_45": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_inadimplencia\sql\faturas_45+.sql",
        "template_xlsx": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\template\TEMPLATE_BASE_ATIVACOES_CANCELAMENTOS.xlsx",
        "img": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\img",
        "history": r"C:\Users\raphael.almeida\Documents\Processos\relatorio_ativacoes_cancelamentos\history",
        "sharepoint": r"C:\Users\raphael.almeida\OneDrive - Grupo Unus\analise de dados - Arquivos em excel"
    }

    def __init__(self):
        # Dataframes
        self.df_ativos = None
        self.df_cancelamentos = None
        
        # Variáveis de Data (Centralizadas)
        self.definir_datas()
        
        # Variáveis de Métricas
        self.metricas = {}

    def definir_datas(self):
        self.hoje = dt.date.today()
        self.dia_semana = self.hoje.weekday()
        
        if self.dia_semana == 0:  # Segunda-feira
            self.yesterday = self.hoje - dt.timedelta(days=1) # Domingo
            self.sabado = self.hoje - dt.timedelta(days=2)
            self.sexta = self.hoje - dt.timedelta(days=3)
            self.resumo_periodo = f"{self.sexta.strftime('%d/%m/%Y')} (sexta) - {self.yesterday.strftime('%d/%m/%Y')} (domingo)"
        else:
            self.yesterday = self.hoje - dt.timedelta(days=1)
            self.sabado = self.yesterday
            self.sexta = self.yesterday
            self.resumo_periodo = self.yesterday.strftime('%d/%m/%Y')
            
        self.yesterday_str = self.yesterday.strftime('%d-%m-%Y')

    def run(self):
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] Iniciando o fluxo de ETL (Ativações e Cancelamentos)...")
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 1/8 - Extraindo dados do Athena...")
        self.extrair_dados()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 2/8 - Processando base de cancelamentos...")
        self.processar_cancelamentos()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 3/8 - Processando faturas em inadimplência...")
        self.processar_inadimplencia()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 4/8 - Calculando as métricas...")
        self.calcular_metricas()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 5/8 - Preenchendo as planilhas do Excel...")
        self.preencher_excel()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 6/8 - Gerando imagem com Html2Image...")
        self.gerar_imagem()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 7/8 - Iniciando automação do WhatsApp (Por favor, não mexa no mouse/teclado)...")
        self.automatizar_whatsapp()
        
        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] 8/8 - Salvando e compartilhando no SharePoint...")
        self.salvar_e_compartilhar()

        print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] Processo finalizado com SUCESSO!")

    @staticmethod
    def format_type(df):
        for col in df.select_dtypes(include=['string']).columns:
            df[col] = df[col].str.upper()
        return df

    def ler_sql(self, path):
        with open(path, 'r', encoding='utf-8') as file:
            return file.read()

    def extrair_dados(self):
        # Ativos
        self.df_ativos = awr.athena.read_sql_query(sql=self.ler_sql(self.PATHS["ativos"]), database='silver', ctas_approach=False)
        self.df_ativos = self.df_ativos.drop(columns=['rn'], errors='ignore')
        self.df_ativos = self.df_ativos.drop_duplicates(subset=['chassi'], keep='first')
        self.df_ativos = self.df_ativos.sort_values(by=['inicio_vig', 'data_ativacao'], ascending=False).reset_index(drop=True)
        self.df_ativos = self.format_type(self.df_ativos)

        # Cancelamentos
        df_canc_int = awr.athena.read_sql_query(sql=self.ler_sql(self.PATHS["cancelamentos_integrais"]), database='silver', ctas_approach=False)
        df_canc_parc = awr.athena.read_sql_query(sql=self.ler_sql(self.PATHS["cancelamentos_parciais"]), database='silver', ctas_approach=False)
        
        df_canc_int = self.format_type(df_canc_int)
        df_canc_parc = self.format_type(df_canc_parc)

        df_canc_parc = df_canc_parc.rename(columns={'data_atualizacao': 'data_cancelamento'})
        df_canc_int['identificador'] = 'INTEGRAL'
        df_canc_parc['identificador'] = 'PARCIAL'

        # Retirando ativos dos parciais
        lista_ativos = self.df_ativos['chassi'].to_list()
        df_canc_parc = df_canc_parc[~df_canc_parc['chassi'].isin(lista_ativos)]

        # Concatenar e limpar
        self.df_cancelamentos = pd.concat([df_canc_int, df_canc_parc], ignore_index=True)
        self.df_cancelamentos['data_cancelamento'] = pd.to_datetime(self.df_cancelamentos['data_cancelamento'], errors='coerce')
        self.df_cancelamentos = self.df_cancelamentos.sort_values(by='data_cancelamento', ascending=False).reset_index(drop=True)
        self.df_cancelamentos = self.df_cancelamentos.drop_duplicates(subset=['chassi'], keep='first')

    def processar_cancelamentos(self):
        def map_franqueado(tem_palavra_chave):
            if pd.isna(tem_palavra_chave):
                return "SIM"
            return "UNIDADES/FRANQUIAS" if tem_palavra_chave else "PARCEIROS/CORRETORES/FTR"

        self.df_cancelamentos['categoria_comercial'] = self.df_cancelamentos['unidade'].str.contains(
            "UNIDADE|MF|FRANQUEADO|MICRO|TS CONSULTORIA|TRANSDESK DIGITAL", na=False, case=False
        ).apply(map_franqueado)
        
        self.df_cancelamentos['data_cancelamento'] = pd.to_datetime(self.df_cancelamentos['data_cancelamento'], errors='coerce').dt.date

    def processar_inadimplencia(self):
        df_faturas = awr.athena.read_sql_query(sql=self.ler_sql(self.PATHS["faturas_45"]), database='silver')
        boletos_pagos = df_faturas.loc[df_faturas['historico'] != 1, 'ponteiro'].tolist()
        df_inadimp = df_faturas[~df_faturas['ponteiro'].isin(boletos_pagos)]
        df_inadimp = df_inadimp[df_inadimp['historico'] == 1]
        df_inadimp = df_inadimp.drop_duplicates(subset=['ponteiro', 'conjunto', 'matricula', 'empresa'], keep='first')

        df_inadimp_atual_45 = df_inadimp[
            ~(
                (df_inadimp['empresa'].isin(['Viavante', 'Stcoop', 'Segtruck'])) &
                (df_inadimp['associado'] == "APROSSIL - ASSOCIACAO DE PROPRIETARIOS DE CAMINHOES DO SUL D")
            )
        ]
        df_inadimp_atual_45 = df_inadimp_atual_45[~df_inadimp_atual_45['associado'].str.contains('TESTE', na=False)]

        self.df_cancelamentos['inadimplente_45+'] = self.df_cancelamentos['conjunto'].isin(df_inadimp_atual_45['conjunto']).map({True: 'SIM', False: 'NÃO'})

    def calcular_metricas(self):
        # Convertendo inicio_vig para datetime para aplicar os filtros
        self.df_ativos['inicio_vig'] = pd.to_datetime(self.df_ativos['inicio_vig'], errors='coerce').dt.date

        if self.dia_semana == 0: 
            ativos_mask = (self.df_ativos['inicio_vig'] >= self.sexta) & (self.df_ativos['inicio_vig'] <= self.yesterday)
            canc_mask = (self.df_cancelamentos['data_cancelamento'] >= self.sexta) & (self.df_cancelamentos['data_cancelamento'] <= self.yesterday)
        else:
            ativos_mask = (self.df_ativos['inicio_vig'] == self.yesterday)
            canc_mask = (self.df_cancelamentos['data_cancelamento'] == self.yesterday)

        parciais_mask = self.df_cancelamentos['identificador'] == 'PARCIAL'
        integrais_mask = self.df_cancelamentos['identificador'] == 'INTEGRAL'
        canc_status_mask = self.df_cancelamentos['status'] == 'CANCELADO'
        fin_status_mask = self.df_cancelamentos['status'] == 'FINALIZADO'

        def get_count(df, empresa, masks):
            combined_mask = (df['empresa'] == empresa)
            for m in masks:
                combined_mask &= m
            return len(df[combined_mask])

        empresas = ["SEGTRUCK", "STCOOP", "VIAVANTE", "TAG"]
        m = self.metricas

        for emp in empresas:
            m[f'ativ_{emp}'] = get_count(self.df_ativos, emp, [ativos_mask])
            m[f'canc_parc_{emp}'] = get_count(self.df_cancelamentos, emp, [canc_mask, parciais_mask])
            m[f'canc_int_{emp}'] = get_count(self.df_cancelamentos, emp, [canc_mask, integrais_mask, canc_status_mask])
            m[f'fin_int_{emp}'] = get_count(self.df_cancelamentos, emp, [canc_mask, integrais_mask, fin_status_mask])

        m['total_ativados'] = len(self.df_ativos[ativos_mask])
        m['total_cancelados'] = len(self.df_cancelamentos[canc_mask])
        
        # Filtros de data para aba base
        m['qtd_ativos_total'] = self.df_ativos[self.df_ativos['inicio_vig'] <= self.yesterday]['chassi'].nunique()
        m['qtd_ativos_sab'] = self.df_ativos[self.df_ativos['inicio_vig'] <= self.sabado]['chassi'].nunique()
        m['qtd_ativos_sex'] = self.df_ativos[self.df_ativos['inicio_vig'] <= self.sexta]['chassi'].nunique()
        
        m['total_ativados_dom'] = len(self.df_ativos[self.df_ativos['inicio_vig'] == self.yesterday])
        m['total_ativados_sab'] = len(self.df_ativos[self.df_ativos['inicio_vig'] == self.sabado])
        m['total_ativados_sex'] = len(self.df_ativos[self.df_ativos['inicio_vig'] == self.sexta])
        
        m['total_canc_dom'] = len(self.df_cancelamentos[self.df_cancelamentos['data_cancelamento'] == self.yesterday])
        m['total_canc_sab'] = len(self.df_cancelamentos[self.df_cancelamentos['data_cancelamento'] == self.sabado])
        m['total_canc_sex'] = len(self.df_cancelamentos[self.df_cancelamentos['data_cancelamento'] == self.sexta])
        
        # Mascaras para escrever no Excel
        self.df_cancelamentos_ontem = self.df_cancelamentos[canc_mask].copy()

    @staticmethod
    def clear_sheet(sheet):
        if sheet.max_row > 1:
            sheet.delete_rows(1, sheet.max_row)

    def preencher_excel(self):
        self.wb = openpyxl.load_workbook(self.PATHS["template_xlsx"])
        w1, w2, w3, w4, w5 = self.wb['ATIVOS'], self.wb['CANCELAMENTOS'], self.wb['BASE'], self.wb['RELATORIO'], self.wb['CANCELAMENTOS ONTEM']

        self.clear_sheet(w1)
        self.clear_sheet(w2)
        self.clear_sheet(w5)

        # Preenchimento W1, W2, W5
        for sheet, df in [(w1, self.df_ativos), (w2, self.df_cancelamentos), (w5, self.df_cancelamentos_ontem)]:
            if not df.empty:
                numeric_cols = df.select_dtypes(include=['number']).columns
                string_cols = df.select_dtypes(include=['object', 'string']).columns
                df[numeric_cols] = df[numeric_cols].fillna(0)
                df[string_cols] = df[string_cols].fillna('N/A')

                for c_idx, col_name in enumerate(df.columns, start=1):
                    sheet.cell(row=1, column=c_idx, value=col_name)
                for r_idx, row in enumerate(df.values, start=2):
                    for c_idx, value in enumerate(row, start=1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

        # Preenchimento W3 (BASE)
        first_empty_row = next((r for r in range(1, w3.max_row + 1) if w3.cell(row=r, column=2).value is None), w3.max_row + 1)

        if self.dia_semana == 0 and first_empty_row >= 3:
            w3[f'B{first_empty_row + 2}'] = self.metricas['qtd_ativos_total']
            w3[f'C{first_empty_row + 2}'] = self.metricas['total_ativados_dom']
            w3[f'D{first_empty_row + 2}'] = self.metricas['total_canc_dom']

            w3[f'B{first_empty_row + 1}'] = self.metricas['qtd_ativos_sab']
            w3[f'C{first_empty_row + 1}'] = self.metricas['total_ativados_sab']
            w3[f'D{first_empty_row + 1}'] = self.metricas['total_canc_sab']

            w3[f'B{first_empty_row}'] = self.metricas['qtd_ativos_sex']
            w3[f'C{first_empty_row}'] = self.metricas['total_ativados_sex']
            w3[f'D{first_empty_row}'] = self.metricas['total_canc_sex']
        else:
            w3[f'B{first_empty_row}'] = self.metricas['qtd_ativos_total']
            w3[f'C{first_empty_row}'] = self.metricas['total_ativados']
            w3[f'D{first_empty_row}'] = self.metricas['total_cancelados']

        # Preenchimento W4 (RELATORIO)
        m = self.metricas
        w4['C2'] = self.resumo_periodo
        w4['C3'] = m['qtd_ativos_total']

        w4['C8'], w4['C9'], w4['C10'], w4['C11'] = m['ativ_SEGTRUCK'], m['ativ_STCOOP'], m['ativ_VIAVANTE'], m['ativ_TAG']
        w4['C12'] = sum([m['ativ_SEGTRUCK'], m['ativ_STCOOP'], m['ativ_VIAVANTE'], m['ativ_TAG']])

        w4['D8'], w4['D9'], w4['D10'], w4['D11'] = m['fin_int_SEGTRUCK'], m['fin_int_STCOOP'], m['fin_int_VIAVANTE'], m['fin_int_TAG']
        w4['D12'] = sum([m['fin_int_SEGTRUCK'], m['fin_int_STCOOP'], m['fin_int_VIAVANTE'], m['fin_int_TAG']])

        w4['E8'], w4['E9'], w4['E10'], w4['E11'] = m['canc_int_SEGTRUCK'], m['canc_int_STCOOP'], m['canc_int_VIAVANTE'], m['canc_int_TAG']
        w4['E12'] = sum([m['canc_int_SEGTRUCK'], m['canc_int_STCOOP'], m['canc_int_VIAVANTE'], m['canc_int_TAG']])

        w4['F8'], w4['F9'], w4['F10'], w4['F11'] = m['canc_parc_SEGTRUCK'], m['canc_parc_STCOOP'], m['canc_parc_VIAVANTE'], m['canc_parc_TAG']
        w4['F12'] = sum([m['canc_parc_SEGTRUCK'], m['canc_parc_STCOOP'], m['canc_parc_VIAVANTE'], m['canc_parc_TAG']])

    def gerar_imagem(self):
        m = self.metricas
        qtd_ativos_fmt = f"{m['qtd_ativos_total']:,}".replace(",", ".")

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <style>
            body {{ font-family: 'Calibri', 'Segoe UI', sans-serif; font-size: 14px; background-color: white; margin: 10px; }}
            table {{ border-collapse: collapse; width: 700px; text-align: center; }}
            th, td {{ border: 1px solid black; padding: 4px 8px; }}
            .bg-gray {{ background-color: #D9D9D9; font-weight: bold; }}
            .bg-orange {{ background-color: #FCE4D6; font-weight: bold; }}
            .bg-yellow {{ background-color: #FFF2CC; font-weight: bold; }}
            .bold {{ font-weight: bold; }}
            .col-empresa {{ width: 15%; text-align: center; }}
            .col-ativadas {{ width: 18%; }}
            .col-fin {{ width: 22%; }}
            .col-canc {{ width: 22%; }}
            .col-ben {{ width: 23%; }}
        </style>
        </head>
        <body>
            <table>
                <tr>
                    <td class="bg-gray col-empresa">DATA</td>
                    <td colspan="4" class="bold">{self.resumo_periodo}</td>
                </tr>
                <tr>
                    <td class="bg-gray">TOTAL ATIVOS</td>
                    <td colspan="4" class="bold">{qtd_ativos_fmt}</td>
                </tr>
            </table><br>
            <table>
                <tr>
                    <td rowspan="3" class="bold bg-gray col-empresa">EMPRESA</td>
                    <td rowspan="3" class="bold bg-gray col-ativadas">PLACAS ATIVADAS</td>
                    <td colspan="3" class="bold bg-gray">PLACAS CANCELADAS</td>
                </tr>
                <tr>
                    <td colspan="2" class="bg-orange">CONJUNTOS CANCELADOS</td>
                    <td rowspan="2" class="bg-yellow">BENEFÍCIOS CANCELADOS</td>
                </tr>
                <tr>
                    <td class="bg-orange col-fin">FINALIZADOS</td>
                    <td class="bg-orange col-canc">CANCELADOS</td>
                </tr>
                <tr>
                    <td class="bg-gray">Segtruck</td>
                    <td>{m['ativ_SEGTRUCK']}</td><td>{m['fin_int_SEGTRUCK']}</td><td>{m['canc_int_SEGTRUCK']}</td><td>{m['canc_parc_SEGTRUCK']}</td>
                </tr>
                <tr>
                    <td class="bg-gray">Stcoop</td>
                    <td>{m['ativ_STCOOP']}</td><td>{m['fin_int_STCOOP']}</td><td>{m['canc_int_STCOOP']}</td><td>{m['canc_parc_STCOOP']}</td>
                </tr>
                <tr>
                    <td class="bg-gray">Viavante</td>
                    <td>{m['ativ_VIAVANTE']}</td><td>{m['fin_int_VIAVANTE']}</td><td>{m['canc_int_VIAVANTE']}</td><td>{m['canc_parc_VIAVANTE']}</td>
                </tr>
                <tr>
                    <td class="bg-gray">Tag</td>
                    <td>{m['ativ_TAG']}</td><td>{m['fin_int_TAG']}</td><td>{m['canc_int_TAG']}</td><td>{m['canc_parc_TAG']}</td>
                </tr>
                <tr>
                    <td class="bold bg-gray">Total</td>
                    <td class="bold">{sum([m['ativ_SEGTRUCK'], m['ativ_STCOOP'], m['ativ_VIAVANTE'], m['ativ_TAG']])}</td>
                    <td class="bold">{sum([m['fin_int_SEGTRUCK'], m['fin_int_STCOOP'], m['fin_int_VIAVANTE'], m['fin_int_TAG']])}</td>
                    <td class="bold">{sum([m['canc_int_SEGTRUCK'], m['canc_int_STCOOP'], m['canc_int_VIAVANTE'], m['canc_int_TAG']])}</td>
                    <td class="bold">{sum([m['canc_parc_SEGTRUCK'], m['canc_parc_STCOOP'], m['canc_parc_VIAVANTE'], m['canc_parc_TAG']])}</td>
                </tr>
            </table>
        </body>
        </html>
        """

        os.makedirs(self.PATHS["img"], exist_ok=True)
        hti = Html2Image(output_path=self.PATHS["img"], size=(750, 350))
        hti.screenshot(html_str=html_content, save_as="relatorio_tabela.png")

    def automatizar_whatsapp(self):
        time.sleep(2)
        pyautogui.hotkey('win', 'e')
        time.sleep(4)
        pyautogui.hotkey('ctrl', 'l') 
        time.sleep(1.5)
        pyautogui.typewrite(self.PATHS["img"], interval=0.05)
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'f') 
        time.sleep(2.5)
        pyautogui.typewrite(r'relatorio')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(1.5)
        pyautogui.press('win')
        time.sleep(1.5)
        pyautogui.typewrite('whatsapp')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(5)
        pyautogui.hotkey('ctrl', 'f') 
        time.sleep(1.5)
        pyautogui.typewrite("ativ")
        time.sleep(1.5)
        pyautogui.press('down')
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2.5)
        pyautogui.typewrite(rf'Bom dia, pessoal! Segue o resultado das placas ativadas e canceladas nas 4 empresas. Refere-se ao(s) dia(s): {self.resumo_periodo}', interval=0.1)
        time.sleep(1.5)
        pyautogui.press('enter')
        time.sleep(1.5)
        
        # exclui arquivo
        pyautogui.hotkey('alt', 'tab')
        time.sleep(1.5)
        pyautogui.press('del')
        time.sleep(1.5)
        pyautogui.hotkey('alt', 'esc')
        time.sleep(1.5)
        pyautogui.hotkey('alt', 'esc')

    def salvar_e_compartilhar(self):
        self.wb.save(self.PATHS["template_xlsx"])
        
        name_file = fr'RELATORIO_ATIVACOES_CANCELAMENTOS_{self.yesterday_str}.xlsx'
        os.makedirs(self.PATHS["history"], exist_ok=True)
        output_full_path = os.path.join(self.PATHS["history"], name_file)
        
        shutil.copy(self.PATHS["template_xlsx"], output_full_path)
        
        name_file_2 = 'RELATORIO_ATIVACOES_CANCELAMENTOS.xlsx'
        full_path_sharepoint = os.path.join(self.PATHS["sharepoint"], name_file_2)
        shutil.copy(output_full_path, full_path_sharepoint)
        
        self.wb.close()

    @classmethod
    def ETL_ativ_cancel(cls):
        relatorio = cls()
        relatorio.run()

if __name__ == "__main__":
    print(f"[{dt.datetime.now().strftime('%H:%M:%S')}] Script acionado. Inicializando classes e carregando dependências...")
    ETL_relat_ativ_cancel.ETL_ativ_cancel()